from pathlib import Path

from openpyxl import load_workbook

from ..abstractions.spreadsheet_reader import SpreadsheetReader


class OpenpyxlReader(SpreadsheetReader):
    def read_header_rows(self, path: Path, sheet: str, max_rows: int = 5) -> list[list]:
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb[sheet]
        rows: list[list] = []
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if i > max_rows:
                break
            rows.append(list(row))
        wb.close()
        return rows

    def read_data_rows(self, path: Path, sheet: str, start_row: int) -> list[list]:
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb[sheet]
        rows: list[list] = []
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if i >= start_row:
                rows.append(list(row))
        wb.close()
        return rows

    def get_sheet_names(self, path: Path) -> list[str]:
        wb = load_workbook(path, read_only=True)
        names = wb.sheetnames
        wb.close()
        return names
